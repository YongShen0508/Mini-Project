import openpyxl
from openpyxl.utils import get_column_letter

def initialize_worksheet(wb, sheet_name):
    ws = wb.active
    if ws.title == "Sheet":
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Set the header row
    ws['B2'] = "Generation / Experiments"
    for i in range(10):
        ws.cell(row=2, column=3+i, value=i+1)

    # Fill B3:B2002 with 1 to 2000
    for i in range(2000):
        ws.cell(row=3+i, column=2, value=i+1)

    # Fill other specified cells
    ws['B2003'] = "Min"
    ws['B2004'] = "Max"
    ws['B2006'] = "Values"
    for i in range(30):
        ws.cell(row=2007+i, column=2, value=i+1)
    ws['B2038'] = "Time"
    
    return ws

def process_file(ws, file_name, column):
    try:
        with open(file_name, 'r') as file:
            current_row = 3

            # Read fitness values
            fitness_values = []  # To store fitness values for later calculations
            for line in file:
                line = line.strip()
                if not line:
                    break
                value = float(line)
                fitness_values.append(value)
                ws.cell(row=current_row, column=column, value=value)
                current_row += 1
            
            col_letter = get_column_letter(column)
            ws.cell(row=2003, column=column, value=f"=MIN({col_letter}3:{col_letter}2002)")
            ws.cell(row=2004, column=column, value=f"=MAX({col_letter}3:{col_letter}2002)")
            
            # Write MIN, MAX, and AVERAGE formulas for row 2002
            ws.cell(row=2002, column=13, value="Min Lowest")
            ws.cell(row=2002, column=14, value="Max Lowest")
            ws.cell(row=2002, column=15, value="Average Fitness")

            # Calculate MIN, MAX, and AVERAGE for row 2002 from columns C to L
            ws.cell(row=2003, column=13, value="=MIN(C2003:L2003)")  # Column M is 13
            ws.cell(row=2003, column=14, value="=MAX(C2003:L2003)")  # Column N is 14
            ws.cell(row=2003, column=15, value="=AVERAGE(C2003:L2003)")  # Column O is 15

            # Skip blank lines to reach chromosome bits
            while True:
                line = file.readline().strip()
                if line:
                    break

            # Read chromosome bits
            current_row = 2007
            while True:
                ws.cell(row=current_row, column=column, value=float(line))
                current_row += 1
                line = file.readline().strip()
                if not line:
                    break

            # Skip blank lines to reach the running time
            while True:
                line = file.readline().strip()
                if line:
                    ws.cell(row=2038, column=column, value=float(line))
                    break
        
            #---------------------------------------------------------------------------------------------#
            #Time calculation

            # Calculate MIN, MAX, and AVERAGE for row 2037 from columns C to L
            ws.cell(row=2037, column=13, value="Min Time")  # Column M is 13
            ws.cell(row=2037, column=14, value="Max Time")  # Column N is 14
            ws.cell(row=2037, column=15, value="Average Time")  # Column O is 15

            # Calculate MIN, MAX, and AVERAGE for row 2038 from columns C to L
            ws.cell(row=2038, column=13, value="=MIN(C2038:L2038)")  # Column M is 13
            ws.cell(row=2038, column=14, value="=MAX(C2038:L2038)")  # Column N is 14
            ws.cell(row=2038, column=15, value="=AVERAGE(C2038:L2038)")  # Column O is 15
            #---------------------------------------------------------------------------------------------#

    except FileNotFoundError:
        print(f"Error: {file_name} not found. Please check if the file is located in the correct directory.")
    except Exception as e:
        print(f"An error occurred while processing {file_name}: {e}")

def save_workbook(wb, workbook_filename, index):
    wb.save(f"{workbook_filename}_{index}.xlsx")

def main(workbook_filename, sheet_info):
    # Initialize variables
    wb_count = 1
    wb = openpyxl.Workbook()
    ws_count = 0
    total_files_processed = 0

    for sheet_name, text_file_prefix, num_files in sheet_info:
        # Create a new workbook if needed
        if ws_count >= 40:
            save_workbook(wb, workbook_filename, wb_count)
            wb_count += 1
            wb = openpyxl.Workbook()
            ws_count = 0

        ws = initialize_worksheet(wb, sheet_name)
        ws_count += 1
        print(f"Starting to process files for {sheet_name}")

        for i in range(1, num_files + 1):
            file_number = total_files_processed + i
            file_name = f"{text_file_prefix}{file_number}.txt"
            column = 3 + ((i - 1) % 10)
            print(f"Processing {file_name} into column {column} of {sheet_name}")
            process_file(ws, file_name, column)

        total_files_processed += num_files

    # Save the last workbook
    if ws_count > 0:
        save_workbook(wb, workbook_filename, wb_count)

if __name__ == "__main__":
    sheet_info = [
    ("GA001", "GAResult", 10),
    ("GA002", "GAResult", 10),
    ("GA003", "GAResult", 10),
    ("GA004", "GAResult", 10),
    ("GA005", "GAResult", 10),
    ("GA006", "GAResult", 10),
    ("GA007", "GAResult", 10),
    ("GA008", "GAResult", 10),
    ("GA009", "GAResult", 10),
    ("GA010", "GAResult", 10),
    ("GA011", "GAResult", 10),
    ("GA012", "GAResult", 10),
    ("GA013", "GAResult", 10),
    ("GA014", "GAResult", 10),
    ("GA015", "GAResult", 10),
    ("GA016", "GAResult", 10),
    ("GA017", "GAResult", 10),
    ("GA018", "GAResult", 10),
    ("GA019", "GAResult", 10),
    ("GA020", "GAResult", 10),
    ("GA021", "GAResult", 10),
    ("GA022", "GAResult", 10),
    ("GA023", "GAResult", 10),
    ("GA024", "GAResult", 10),
    ("GA025", "GAResult", 10),
    ("GA026", "GAResult", 10),
    ("GA027", "GAResult", 10),
    ("GA028", "GAResult", 10),
    ("GA029", "GAResult", 10),
    ("GA030", "GAResult", 10),
    ("GA031", "GAResult", 10),
    ("GA032", "GAResult", 10),
    ("GA033", "GAResult", 10),
    ("GA034", "GAResult", 10),
    ("GA035", "GAResult", 10),
    ("GA036", "GAResult", 10),
    ("GA037", "GAResult", 10),
    ("GA038", "GAResult", 10),
    ("GA039", "GAResult", 10),
    ("GA040", "GAResult", 10),
    ("GA041", "GAResult", 10),
    ("GA042", "GAResult", 10),
    ("GA043", "GAResult", 10),
    ("GA044", "GAResult", 10),
    ("GA045", "GAResult", 10),
    ("GA046", "GAResult", 10),
    ("GA047", "GAResult", 10),
    ("GA048", "GAResult", 10),
    ("GA049", "GAResult", 10),
    ("GA050", "GAResult", 10),
    ("GA051", "GAResult", 10),
    ("GA052", "GAResult", 10),
    ("GA053", "GAResult", 10),
    ("GA054", "GAResult", 10),
    ("GA055", "GAResult", 10),
    ("GA056", "GAResult", 10),
    ("GA057", "GAResult", 10),
    ("GA058", "GAResult", 10),
    ("GA059", "GAResult", 10),
    ("GA060", "GAResult", 10),
    ("GA061", "GAResult", 10),
    ("GA062", "GAResult", 10),
    ("GA063", "GAResult", 10),
    ("GA064", "GAResult", 10),
    ("GA065", "GAResult", 10),
    ("GA066", "GAResult", 10),
    ("GA067", "GAResult", 10),
    ("GA068", "GAResult", 10),
    ("GA069", "GAResult", 10),
    ("GA070", "GAResult", 10),
    ("GA071", "GAResult", 10),
    ("GA072", "GAResult", 10),
    ("GA073", "GAResult", 10),
    ("GA074", "GAResult", 10),
    ("GA075", "GAResult", 10),
    ("GA076", "GAResult", 10),
    ("GA077", "GAResult", 10),
    ("GA078", "GAResult", 10),
    ("GA079", "GAResult", 10),
    ("GA080", "GAResult", 10),
    ("GA081", "GAResult", 10),
    ("GA082", "GAResult", 10),
    ("GA083", "GAResult", 10),
    ("GA084", "GAResult", 10),
    ("GA085", "GAResult", 10),
    ("GA086", "GAResult", 10),
    ("GA087", "GAResult", 10),
    ("GA088", "GAResult", 10),
    ("GA089", "GAResult", 10),
    ("GA090", "GAResult", 10),
    ("GA091", "GAResult", 10),
    ("GA092", "GAResult", 10),
    ("GA093", "GAResult", 10),
    ("GA094", "GAResult", 10),
    ("GA095", "GAResult", 10),
    ("GA096", "GAResult", 10),
    ("GA097", "GAResult", 10),
    ("GA098", "GAResult", 10),
    ("GA099", "GAResult", 10),
    ("GA100", "GAResult", 10),
    ("GA101", "GAResult", 10),
    ("GA102", "GAResult", 10),
    ("GA103", "GAResult", 10),
    ("GA104", "GAResult", 10),
    ("GA105", "GAResult", 10),
    ("GA106", "GAResult", 10),
    ("GA107", "GAResult", 10),
    ("GA108", "GAResult", 10),
    ("GA109", "GAResult", 10),
    ("GA110", "GAResult", 10),
    ("GA111", "GAResult", 10),
    ("GA112", "GAResult", 10),
    ("GA113", "GAResult", 10),
    ("GA114", "GAResult", 10),
    ("GA115", "GAResult", 10),
    ("GA116", "GAResult", 10),
    ("GA117", "GAResult", 10),
    ("GA118", "GAResult", 10),
    ("GA119", "GAResult", 10),
    ("GA120", "GAResult", 10),
    ("GA121", "GAResult", 10),
    ("GA122", "GAResult", 10),
    ("GA123", "GAResult", 10),
    ("GA124", "GAResult", 10),
    ("GA125", "GAResult", 10),
    ("GA126", "GAResult", 10),
    ("GA127", "GAResult", 10),
    ("GA128", "GAResult", 10),
    ("GA129", "GAResult", 10),
    ("GA130", "GAResult", 10),
    ("GA131", "GAResult", 10),
    ("GA132", "GAResult", 10),
    ("GA133", "GAResult", 10),
    ("GA134", "GAResult", 10),
    ("GA135", "GAResult", 10),
    ("GA136", "GAResult", 10),
    ("GA137", "GAResult", 10),
    ("GA138", "GAResult", 10),
    ("GA139", "GAResult", 10),
    ("GA140", "GAResult", 10),
    ("GA141", "GAResult", 10),
    ("GA142", "GAResult", 10),
    ("GA143", "GAResult", 10),
    ("GA144", "GAResult", 10),
    ("GA145", "GAResult", 10),
    ("GA146", "GAResult", 10),
    ("GA147", "GAResult", 10),
    ("GA148", "GAResult", 10),
    ("GA149", "GAResult", 10),
    ("GA150", "GAResult", 10),
    ("GA151", "GAResult", 10),
    ("GA152", "GAResult", 10),
    ("GA153", "GAResult", 10),
    ("GA154", "GAResult", 10),
    ("GA155", "GAResult", 10),
    ("GA156", "GAResult", 10),
    ("GA157", "GAResult", 10),
    ("GA158", "GAResult", 10),
    ("GA159", "GAResult", 10),
    ("GA160", "GAResult", 10),
    ("GA161", "GAResult", 10),
    ("GA162", "GAResult", 10),
    ("GA163", "GAResult", 10),
    ("GA164", "GAResult", 10),
    ("GA165", "GAResult", 10),
    ("GA166", "GAResult", 10),
    ("GA167", "GAResult", 10),
    ("GA168", "GAResult", 10),
    ("GA169", "GAResult", 10),
    ("GA170", "GAResult", 10),
    ("GA171", "GAResult", 10),
    ("GA172", "GAResult", 10),
    ("GA173", "GAResult", 10),
    ("GA174", "GAResult", 10),
    ("GA175", "GAResult", 10),
    ("GA176", "GAResult", 10),
    ("GA177", "GAResult", 10),
    ("GA178", "GAResult", 10),
    ("GA179", "GAResult", 10),
    ("GA180", "GAResult", 10),
    ("GA181", "GAResult", 10),
    ("GA182", "GAResult", 10),
    ("GA183", "GAResult", 10),
    ("GA184", "GAResult", 10),
    ("GA185", "GAResult", 10),
    ("GA186", "GAResult", 10),
    ("GA187", "GAResult", 10),
    ("GA188", "GAResult", 10),
    ("GA189", "GAResult", 10),
    ("GA190", "GAResult", 10),
    ("GA191", "GAResult", 10),
    ("GA192", "GAResult", 10),
    ("GA193", "GAResult", 10),
    ("GA194", "GAResult", 10),
    ("GA195", "GAResult", 10),
    ("GA196", "GAResult", 10),
    ("GA197", "GAResult", 10),
    ("GA198", "GAResult", 10),
    ("GA199", "GAResult", 10),
    ("GA200", "GAResult", 10),
    ("GA201", "GAResult", 10),
    ("GA202", "GAResult", 10),
    ("GA203", "GAResult", 10),
    ("GA204", "GAResult", 10),
    ("GA205", "GAResult", 10),
    ("GA206", "GAResult", 10),
    ("GA207", "GAResult", 10),
    ("GA208", "GAResult", 10),
    ("GA209", "GAResult", 10),
    ("GA210", "GAResult", 10),
    ("GA211", "GAResult", 10),
    ("GA212", "GAResult", 10),
    ("GA213", "GAResult", 10),
    ("GA214", "GAResult", 10),
    ("GA215", "GAResult", 10),
    ("GA216", "GAResult", 10),
    ("GA217", "GAResult", 10),
    ("GA218", "GAResult", 10),
    ("GA219", "GAResult", 10),
    ("GA220", "GAResult", 10),
    ("GA221", "GAResult", 10),
    ("GA222", "GAResult", 10),
    ("GA223", "GAResult", 10),
    ("GA224", "GAResult", 10),
    ("GA225", "GAResult", 10),
    ("GA226", "GAResult", 10),
    ("GA227", "GAResult", 10),
    ("GA228", "GAResult", 10),
    ("GA229", "GAResult", 10),
    ("GA230", "GAResult", 10),
    ("GA231", "GAResult", 10),
    ("GA232", "GAResult", 10),
    ("GA233", "GAResult", 10),
    ("GA234", "GAResult", 10),
    ("GA235", "GAResult", 10),
    ("GA236", "GAResult", 10),
    ("GA237", "GAResult", 10),
    ("GA238", "GAResult", 10),
    ("GA239", "GAResult", 10),
    ("GA240", "GAResult", 10),
    ("GA241", "GAResult", 10),
    ("GA242", "GAResult", 10),
    ("GA243", "GAResult", 10),
    ("GA244", "GAResult", 10),
    ("GA245", "GAResult", 10),
    ("GA246", "GAResult", 10),
    ("GA247", "GAResult", 10),
    ("GA248", "GAResult", 10),
    ("GA249", "GAResult", 10),
    ("GA250", "GAResult", 10),
    ("GA251", "GAResult", 10),
    ("GA252", "GAResult", 10),
    ("GA253", "GAResult", 10),
    ("GA254", "GAResult", 10),
    ("GA255", "GAResult", 10),
    ("GA256", "GAResult", 10),
    ]
    temp_list = []

    for prefix, result, count in sheet_info:
        for i in range(1, 11):
            temp_list.append((f"{prefix}_{i}", result, count))

    main("GA_Results", temp_list)